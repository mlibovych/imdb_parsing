from bs4 import BeautifulSoup
import lxml
import requests
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl import load_workbook
import os

from dbclass import Database

database = Database()

if os.name == 'posix':
    delimeter = '/'
else:
    delimeter = '\\'

imdb_url = 'https://www.imdb.com/'
exact_matches_url_parts = ['https://www.imdb.com/find?q=', 'film', '&s=tt&exact=true&ref_=fn_tt_ex']

captions_table = 'captions.xlsx'

if not os.path.exists(captions_table):
    wb = openpyxl.Workbook()
else:
    wb = load_workbook(captions_table)

DONE = 1
NOT_DONE = 0

def get_img(film_dir, src):
    print('Start image downloading')

    request = requests.get(src)
    ext = os.path.splitext(src)[1]

    with open(film_dir + delimeter + 'poster' + ext, 'wb') as f:
        f.write(request.content)

    print('done')

def wait_for_js_loads_video(player):
    # open browser
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    browser = webdriver.Chrome(os.path.dirname(os.path.abspath(__file__)) + delimeter + 'chromedriver',
                               options=chrome_options)
    browser.get(imdb_url + player.find('a')['href'])

    # wait for JS load video
    for i in range(100):
        try:
            element = browser.find_element_by_tag_name('video')
            time.sleep(1)
            break
        except NoSuchElementException:
            time.sleep(0.1)

    # get page code
    html = browser.page_source

    # close browser
    browser.close()

    return html


def get_video_link(player):
    # get page code
    html = wait_for_js_loads_video(player)

    # get link
    soup = BeautifulSoup(html, 'lxml')
    player = soup.find('div', class_='video-player__video')
    video = player.find('video')

    if video is None:
        return None

    link = video['src']

    return link


def get_video(film_dir, link):
    print('Start video downloading')

    request = requests.get(link, stream=True)

    with open(film_dir + delimeter + 'trailer' + '.mp4', 'wb') as f:
        for chunk in request.iter_content(chunk_size=1024):
            if chunk:  # filter out keep-alive new chunks
                f.write(chunk)

    print('done')


def find_film(film_title):
    film_title = film_title.replace('&', '%26')
    exact_matches_url_parts[1] = film_title.replace(' ', '+')
    exact_matches_url = ''.join(exact_matches_url_parts)
    print('search', exact_matches_url)

    request = requests.get(exact_matches_url)
    soup = BeautifulSoup(request.content, 'lxml')
    find_section = soup.find(class_='findSection')

    if find_section is not None:
        find_list = find_section.find('table', class_='findList')
        find_result = find_list.find('tr', class_='findResult')
        link = find_result.find('td', class_='result_text').find('a')['href']

        return imdb_url + link
    else:
        return None


def get_section_links(article, section_title):
    if article is None:
        return None

    section_header = article.find(class_='inline', text=section_title)

    if section_header is None:
        return None

    elements = []

    for sib in section_header.next_siblings:
        if sib.name == 'a':
            elements.append(sib.text)

    return elements


def search_keywords(link):
    request = requests.get(imdb_url + link)
    soup = BeautifulSoup(request.content, 'lxml')
    table = soup.find('table', class_='dataTable evenWidthTable2Col')

    if table is None:
        return None

    rows = table.find_all('tr');
    keywords = []
    count = 0

    for row in rows:
        cells = row.find_all('td')

        for cell in cells:
            href = cell.find('a')

            if href is not None:
                keyword = cell.find('a').text
                keywords.append(keyword)
                count += 1
        if count >= 40:
            break

    return keywords


def get_keywords(story_lane):
    if story_lane is None:
        return None

    keywords_header = story_lane.find(class_='inline', text='Plot Keywords:')
    if keywords_header is None:
        return None

    all_keywords = keywords_header.find_next_sibling('nobr')

    if all_keywords is None:
        return None

    link = all_keywords.find('a')['href']

    return search_keywords(link)


def get_reviews(film_link):
    request = requests.get(film_link + 'externalreviews')
    soup = BeautifulSoup(request.content, 'lxml')
    table = soup.find('div', id='external_reviews_content')

    if table is None:
        return None

    empty = table.find('div', id='no_content')

    if empty is not None:
        return None

    review_links = table.find_all('a')
    result_links = []
    count = 0

    for link in review_links:
        try:
            review_request = requests.get(imdb_url + link['href'], timeout=3)
            result_links.append(link.text + " " + review_request.url)
        except Exception:
            continue
        count += 1
        if count > 2:
            break

    return result_links


def download_media(title_overview, film_title):
    # download poster
    img = title_overview.find('div', class_='slate_wrapper').find('div', class_='poster').find('img')

    get_img(film_title, img['src'])

    # download trailer
    video = title_overview.find('div', class_='slate_wrapper').find('div', class_='slate')
    video_link = get_video_link(video)

    if video_link is not None:
        get_video(film_title, video_link)
    else:
        print('Can not download video')


def get_film_data(film_id, film_title, film_link):
    # prepare links
    request = requests.get(film_link)
    soup = BeautifulSoup(request.content, 'lxml')
    content = soup.find('div', id='wrapper').find('div', id='content-2-wide')
    top = content.find('div', id='main_top', class_='main')
    bottom = content.find('div', id='main_bottom', class_='main')
    title_overview = top.find('div', class_='title-overview')

    # sections
    story_lane = bottom.find('div', id='titleStoryLine')
    details = bottom.find('div', id='titleDetails')

    # get media
    # download_media(title_overview, film_title)

    # get general data

    # director
    plot_summary = title_overview.find('div', class_='plot_summary_wrapper')
    if plot_summary is None:
        director_name = 'None'
    else:
        credit_summary = plot_summary.find('div', class_='credit_summary_item')
        if credit_summary is None:
            director_name = 'None'
        else:
            director_name = credit_summary.find('a').text

    # genres
    genres_list = get_section_links(story_lane, 'Genres:')
    if genres_list is None:
        genres = 'None'
    else:
        genres = ", ".join(genres_list)

    # country
    countries_list = get_section_links(details, 'Country:')
    if countries_list is None:
        countries = 'None'
    else:
        countries = ", ".join(countries_list)

    # prodaction co
    companies_list = get_section_links(details, 'Production Co:')
    if companies_list is None:
        companies = 'None'
    else:
        companies = ", ".join(companies_list)

    # release date
    if details is None:
        date_header = None
    else:
        date_header = details.find(class_='inline', text='Release Date:')
    if date_header is None:
        date = 'None'
    else:
        date = date_header.next_element.next_element

    # keywords
    keywords_list = get_keywords(story_lane)
    if keywords_list is None:
        keywords = 'None'
    else:
        keywords = ", ".join(keywords_list)

    # reviews
    reviews_list = get_reviews(film_link)
    if reviews_list is None:
        reviews = 'None'
    else:
        reviews = ", ".join(reviews_list)


    # save in DB
    data_id = database.insert_film_data(film_id, director_name, genres,
                                        countries, companies,
                                        date, keywords, reviews)[0]

    print('start xls')
    # save in xlsx
    write_in_xlsx(data_id, film_id, director_name,
                  genres, countries, companies,
                  date, keywords, reviews)


    # create txt file
    # with open(film_title + delimeter + film_title + '.txt', 'w') as f:
    #     f.write(film_title + '\n\n')
    #
    #     f.write('Director: ' + director_name + '\n\n')
    #
    #     f.write('Genres:' + '\n')
    #     for genre in genres:
    #         f.write('\t' + genre.strip() + '\n')
    #     f.write('\n')
    #
    #     f.write('Country:' + '\n')
    #     for country in countries:
    #         f.write('\t' + country.strip() + '\n')
    #     f.write('\n')
    #
    #     f.write('Prodaction Co:' + '\n')
    #     for company in companies:
    #         f.write('\t' + company.strip() + '\n')
    #     f.write('\n')
    #
    #     if date is not None:
    #         f.write('Release Date: ' + date.strip() + '\n\n')
    #
    #     f.write('Keywords:' + '\n')
    #     for keyword in keywords:
    #         f.write('\t' + keyword.strip() + '\n')
    #     f.write('\n')
    #
    #     f.write('Reviews:' + '\n')
    #     for review in reviews:
    #         f.write('\t' + review + '\n')
    #     f.write('\n')


def proccess_film(film_id, film_title):
    film_title = film_title.replace('/', '-')

    # if not os.path.exists(film_title):
    #     os.mkdir(film_title)

    link = find_film(film_title)

    if link is None:
        print('No exact matches')
    else:
        get_film_data(film_id, film_title, link)

    # database.update_film(film_id, DONE)
    print('Done')


def write_in_xlsx(id, film_id, director, genres, country, production_co,
                  release_date, keywords, reviews):
    sheets = wb.sheetnames

    if 'films_data' not in sheets:
        # create sheet & insert headers
        wb.create_sheet(title='films_data', index=len(sheets))
        data_sheet = wb['films_data']

        headers = ['id', 'film_id', 'director', 'genres', 'country', 'production_co',
                   'release_date', 'keywords', 'reviews']
        col = 1
        for header in headers:
            data_sheet.cell(row=1, column=col).value = header
            col += 1

    data_sheet = wb['films_data']
    row = data_sheet.max_row + 1

    data_sheet.cell(row=row, column=1).value = id
    data_sheet.cell(row=row, column=2).value = film_id
    data_sheet.cell(row=row, column=3).value = director
    data_sheet.cell(row=row, column=4).value = genres
    data_sheet.cell(row=row, column=5).value = country
    data_sheet.cell(row=row, column=6).value = production_co
    data_sheet.cell(row=row, column=7).value = release_date
    data_sheet.cell(row=row, column=8).value = keywords
    data_sheet.cell(row=row, column=9).value = reviews

    # wb.save(captions_table)


def set_empty_reviews():
    data_sheet = wb['films_data']

    for row in range(1, data_sheet.max_row):
        if data_sheet.cell(row=row, column=9).value.strip() == 'External Reviews submission guide.':
            data_sheet.cell(row=row, column=9).value = 'None'

    wb.save(captions_table)


def remove_empty_sheet():
    std = wb['Sheet']

    wb.remove_sheet(std)
    wb.save(captions_table)


def main():
    films = database.get_not_proceeded_films()
    ids = []

    for id, film in films:
        print(id, film)
        # parse film data
        proccess_film(id, film)

        ids.append(id)
        if len(ids) >= 20:
            # update DB data
            for film_id in ids:
                database.update_film(film_id, DONE)

            # clear id list
            ids.clear()
            # save xlsx file
            wb.save(captions_table)
            print("Saved")
            time.sleep(3)

    wb.save(captions_table)


main()

